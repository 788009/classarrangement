import sys
import os
import json
import openpyxl
import pandas as pd
from itertools import groupby
from datetime import datetime
import traceback
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, \
     QLabel, QPushButton, QFileDialog, QTableWidget, QTableWidgetItem, \
     QLineEdit, QHBoxLayout, QHeaderView, QComboBox, QScrollArea, \
     QListWidget, QListView, QAbstractItemView, QCheckBox, QDialog, QVBoxLayout,\
     QMenu, QAction, QMessageBox, QTextEdit
from PyQt5.QtCore import QDir, Qt, QEvent

class MultiSelectCellWidget(QPushButton):
    def __init__(self, row, table_widget, mainWidget, parent=None):
        super().__init__(parent)
        self.menu = QMenu(self)
        self.checkbox_actions = []
        self.init_menu()
        self.row = row
        self.table_widget = table_widget
        self.mainWidget = mainWidget

        self.setText("点击选择")
        self.clicked.connect(self.show_menu)

    def init_menu(self):
        self.subjects = [
            '物理、化学、生物',
            '物理、化学、地理',
            '物理、化学、政治',
            '物理、生物、地理',
            '物理、生物、政治',
            '物理、地理、政治',
            '历史、化学、生物',
            '历史、化学、地理',
            '历史、化学、政治',
            '历史、生物、地理',
            '历史、生物、政治',
            '历史、地理、政治'
        ]

        for subject in self.subjects:
            action = QAction(subject, self.menu, checkable=True)
            self.menu.addAction(action)
            self.checkbox_actions.append(action)

    def show_menu(self):
        for action in self.checkbox_actions:
            action.triggered.connect(lambda checked, action=action: self.update_button_text())

        pos = self.mapToGlobal(self.rect().bottomLeft())
        self.menu.exec_(pos)

    def update_button_text(self):
        selected_items = [action.text() for action in self.checkbox_actions if action.isChecked()]
        self.setText('\n'.join(selected_items))
        self.adjust_row_height()

    def setText(self, text):
        super().setText(text)
        selected_subjects = text.split('\n')
        for action in self.checkbox_actions:
            action.setChecked(action.text() in selected_subjects)
        self.adjust_row_height()

    def adjust_row_height(self):
        row_height_per_line = 20  # 每行的高度，可以根据需要调整
        lines = self.text().count('\n') + 1
        total_height = lines * row_height_per_line
        if lines == 1:
            total_height = 30

        self.table_widget.setRowHeight(self.row, total_height)
        self.mainWidget.adjustTableHeight(self.table_widget)

class ProgressDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.initUI()

    def initUI(self):
        self.setWindowTitle('进度')
        self.setGeometry(100, 100, 400, 300)
        
        self.textEdit = QTextEdit(self)
        self.textEdit.setReadOnly(True)

        layout = QVBoxLayout()
        layout.addWidget(self.textEdit)
        self.setLayout(layout)

    def updateText(self, msg):
        self.textEdit.append(msg)
        QApplication.processEvents()

class MainWidget(QWidget):
    def __init__(self):
        super().__init__()

        self.setWindowTitle('分班')
        self.setGeometry(100, 100, 800, 600)
        self.setStyleSheet("""
            QWidget {
                font-family: Arial;
                font-size: 14px;
            }
            QPushButton#normal-button {
                background-color: #3498db;
                color: white;
                border-radius: 5px;
                padding: 10px;
                border: none;
                transition: background-color 0.3s;
            }
            QPushButton#normal-button:hover {
                background-color: #0e90d2;
            }
            QPushButton#submit-button {
                background-color: #f22c3d;
                color: white;
                border-radius: 5px;
                padding: 10px;
                border: none;
                transition: background-color 0.3s;
            }
            QPushButton#submit-button:hover {
                background-color: #ff4055;
            }
            QLabel {
                margin: 5px;
            }
            QTableWidget {
                gridline-color: #ddd;
                background-color: #f9f9f9;
                border: 1px solid #ddd;
            }
            QTableWidget QTableCornerButton::section {
                background-color: #f1f1f1;
                border: 1px solid #ddd;
            }
            QHeaderView::section {
                background-color: #f1f1f1;
                border: 1px solid #ddd;
                padding: 5px;
            }
            QLineEdit {
                border: 1px solid #ddd;
                border-radius: 3px;
                padding: 5px;
            }
            QPushButton#delete-button {
                background-color: #3498db;
                color: white;
                border-radius: 0px;
                padding: 0px;
                border: none;
                transition: background-color 0.3s;
            }
            QPushButton#delete-button:hover {
                background-color: #0e90d2;
            }
            QComboBox::item { padding: 5px; }
            QComboBox::item::text { align: center; }
        """)

        # 创建一个 QScrollArea 并设置其为滚动窗口
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
 
        # 创建一个 QWidget 作为滚动区域的内容
        self.scroll_content = QWidget(self.scroll_area)
        self.scroll_area.setWidget(self.scroll_content)
 
        # 创建布局并添加一些内容
        self.layout = QVBoxLayout(self.scroll_content)
        
        #成绩文件夹
        self.select_folder_button = QPushButton('选择成绩文件夹')
        self.select_folder_button.setObjectName('normal-button')
        self.select_folder_button.clicked.connect(self.selectGradeFolder)
        self.layout.addWidget(self.select_folder_button)

        self.gradeLabel = QLabel('成绩文件夹（包含若干考试数据表格）：')
        self.layout.addWidget(self.gradeLabel)
        
        self.gradeTable = QTableWidget(self)
        self.gradeTable.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.layout.addWidget(self.gradeTable)
        
        #选科表格
        self.subButton = QPushButton('选择选科表格')
        self.subButton.setObjectName('normal-button')
        self.subButton.clicked.connect(self.selectSubject)
        self.layout.addWidget(self.subButton)

        self.subNoticeLabel = QLabel('格式：姓名 选科1 选科2 选科3')
        self.layout.addWidget(self.subNoticeLabel)

        self.subLabel = QLabel('选科表格：')
        self.layout.addWidget(self.subLabel)
        
        #性别表格
        self.genderButton = QPushButton('选择性别表格')
        self.genderButton.setObjectName('normal-button')
        self.genderButton.clicked.connect(self.selectGender)
        self.layout.addWidget(self.genderButton)

        self.genderNoticeLabel = QLabel('格式：姓名 性别')
        self.layout.addWidget(self.genderNoticeLabel)

        self.genderLabel = QLabel('性别表格：')
        self.layout.addWidget(self.genderLabel)
        
        #班型
        self.cTypeLabel = QLabel('''班型：
班级序号之间用英文逗号隔开，支持输入范围，示例：1-3,5
“选科”可多选''')
        self.layout.addWidget(self.cTypeLabel)

        # 添加新的 QTableWidget 用于选择班型和班级序号
        self.classTable = QTableWidget(self.scroll_content)
        self.classTable.setColumnCount(4)
        self.classTable.setHorizontalHeaderLabels(['选科', '班型', '班级序号', '删除'])
        self.classTable.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.populateClassTable()
        self.layout.addWidget(self.classTable)
        
        # 添加“添加班型”按钮
        self.addClassButton = QPushButton('添加班型')
        self.addClassButton.setObjectName('normal-button')
        self.addClassButton.clicked.connect(self.addClassRow)
        self.layout.addWidget(self.addClassButton)

        self.grButton = QPushButton('生成分班表格')
        self.grButton.setObjectName('submit-button')
        self.grButton.clicked.connect(self.generateResult)
        self.layout.addWidget(self.grButton)

        # 添加存草稿和读取草稿按钮
        self.saveDraftButton = QPushButton('存草稿')
        self.saveDraftButton.setObjectName('normal-button')
        self.saveDraftButton.clicked.connect(self.saveDraft)
        self.layout.addWidget(self.saveDraftButton)

        self.loadDraftButton = QPushButton('读取草稿')
        self.loadDraftButton.setObjectName('normal-button')
        self.loadDraftButton.clicked.connect(self.loadDraft)
        self.layout.addWidget(self.loadDraftButton)

        self.layout.setSpacing(15)
        self.layout.setContentsMargins(20, 20, 20, 20)

        # 设置窗口的布局为 QScrollArea
        main_layout = QVBoxLayout(self)
        main_layout.addWidget(self.scroll_area)

        # 调整表格高度以适应所有行
        self.adjustTableHeight(self.classTable)
        self.adjustTableHeight(self.gradeTable)

    def selectGradeFolder(self):
        folder_path = QFileDialog.getExistingDirectory(self, '选择成绩文件夹')
        if folder_path:
            self.displayFiles(folder_path)
            self.gradeLabel.setText(f'成绩文件夹：{folder_path}')

    def selectSubject(self):
        file_path, _ = QFileDialog.getOpenFileName(self, '选择选科表格', '', 'Excel files (*.xlsx *.xls)')
        self.subLabel.setText(f'选科表格：{file_path}')

    def selectGender(self):
        file_path, _ = QFileDialog.getOpenFileName(self, '选择性别表格', '', 'Excel files (*.xlsx *.xls)')
        self.genderLabel.setText(f'性别表格：{file_path}')        

    def displayFiles(self, folder_path):
        files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx') or f.endswith('.xls')]
        self.gradeTable.setRowCount(len(files))
        self.gradeTable.setColumnCount(5)
        self.gradeTable.setHorizontalHeaderLabels(['考试名', '权重 / %', '姓名所在列', '分数所在列', '排名所在列'])

        for i, file_name in enumerate(files):
            exam_name_item = QTableWidgetItem('.'.join(file_name.split('.')[:-1]))
            self.gradeTable.setItem(i, 0, exam_name_item)
            self.addLineEdit(i, 1)
            self.addLineEdit(i, 2, 'A')
            self.addLineEdit(i, 3, 'B')
            self.addLineEdit(i, 4, 'C')

        self.gradeTable.resizeColumnsToContents()
        self.adjustTableHeight(self.gradeTable)

    def addLineEdit(self, row, col, value=''):
        line_edit = QLineEdit()
        line_edit.setAlignment(Qt.AlignCenter)
        line_edit.setText(value)
        self.gradeTable.setCellWidget(row, col, line_edit)

    def populateClassTable(self):
        subjects = ['物理、化学、生物', '物理、化学、生物', '历史、地理、政治', '历史、地理、政治']
        ctypes = ['实验班', '普通班', '实验班', '普通班']
        self.classTable.setRowCount(len(subjects))
        for i, subject in enumerate(subjects):
            self.populateRow(i, subject, ctypes[i])

    def populateRow(self, row, subject, ctype):
        multi_select_widget = MultiSelectCellWidget(row, self.classTable, self)
        multi_select_widget.setText(subject)
        self.classTable.setCellWidget(row, 0, multi_select_widget)

        combo_box_class = QComboBox()
        combo_box_class.installEventFilter(self)
        combo_box_class.addItems(['普通班', '实验班'])
        combo_box_class.setCurrentText(ctype)
        self.classTable.setCellWidget(row, 1, combo_box_class)

        item_class_number = QTableWidgetItem('')
        item_class_number.setTextAlignment(Qt.AlignCenter)
        self.classTable.setItem(row, 2, item_class_number)

        delete_button = QPushButton('删除')
        delete_button.setObjectName('delete-button')
        delete_button.clicked.connect(lambda _, r=row: self.removeClassRow(r))
        self.classTable.setCellWidget(row, 3, delete_button)

        self.adjustTableHeight(self.classTable)

    def addClassRow(self):
        current_rows = self.classTable.rowCount()
        self.classTable.setRowCount(current_rows + 1)
        self.populateRow(current_rows, '', '')
        self.adjustTableHeight(self.classTable)

    def removeClassRow(self, row):
        self.classTable.removeRow(row)
        self.adjustTableHeight(self.classTable)

    def adjustTableHeight(self, table):
        table_height = table.verticalHeader().length() + table.horizontalHeader().height()
        table.setFixedHeight(table_height)

    def eventFilter(self, obj, event):
        if isinstance(obj, QComboBox) and event.type() == QEvent.Wheel:
            # 如果事件类型是滚轮事件且对象是 QComboBox，直接返回True以阻止事件传递
            return True
        return super().eventFilter(obj, event)

    def saveDraft(self):
        data = {
            'grade_folder': self.gradeLabel.text(),
            'subject_file': self.subLabel.text(),
            'gender_file': self.genderLabel.text(),
            'class_table': [],
            'grade_table': []
        }

        for row in range(self.classTable.rowCount()):
            subject_widget = self.classTable.cellWidget(row, 0)
            subject = subject_widget.text() if subject_widget else ''

            class_type_widget = self.classTable.cellWidget(row, 1)
            class_type = class_type_widget.currentText() if class_type_widget else ''

            class_number_item = self.classTable.item(row, 2)
            class_number = class_number_item.text() if class_number_item else ''

            data['class_table'].append({
                'subject': subject,
                'class_type': class_type,
                'class_number': class_number
            })

        for row in range(self.gradeTable.rowCount()):
            file_item = self.gradeTable.item(row, 0)
            file_name = file_item.text() if file_item else ''

            weight_widget = self.gradeTable.cellWidget(row, 1)
            weight = weight_widget.text() if weight_widget else ''

            data['grade_table'].append({
                'file_name': file_name,
                'weight': weight
            })

        file_path, _ = QFileDialog.getSaveFileName(self, '保存草稿', '', 'JSON files (*.json)')
        if file_path:
            with open(file_path, 'w') as file:
                json.dump(data, file)

    def loadDraft(self):
        file_path, _ = QFileDialog.getOpenFileName(self, '读取草稿', '', 'JSON files (*.json)')
        if file_path:
            with open(file_path, 'r') as file:
                data = json.load(file)

            self.gradeLabel.setText(data.get('grade_folder', ''))
            self.subLabel.setText(data.get('subject_file', ''))
            self.genderLabel.setText(data.get('gender_file', ''))

            # 如果成绩文件夹存在，则显示文件
            grade_folder = data.get('grade_folder', '')
            if grade_folder:
                self.displayFiles(grade_folder.split('：')[-1])

            # 读取并设置权重
            grade_table_data = data.get('grade_table', [])
            for row, grade_data in enumerate(grade_table_data):
                file_name = grade_data.get('file_name', '')
                weight = grade_data.get('weight', '')

                weight_widget = QLineEdit()
                weight_widget.setAlignment(Qt.AlignCenter)
                weight_widget.setText(weight)
                self.gradeTable.setCellWidget(row, 1, weight_widget)

            # 读取班级表格
            class_table_data = data.get('class_table', [])
            self.classTable.setRowCount(len(class_table_data))
            for row, class_data in enumerate(class_table_data):
                subject_widget = MultiSelectCellWidget(row, self.classTable, self)
                subject_widget.setText(class_data.get('subject', ''))
                self.classTable.setCellWidget(row, 0, subject_widget)

                class_type_widget = QComboBox()
                class_type_widget.installEventFilter(self)
                class_type_widget.addItems(['实验班', '普通班'])
                class_type_widget.setCurrentText(class_data.get('class_type', ''))
                self.classTable.setCellWidget(row, 1, class_type_widget)

                class_number_item = QTableWidgetItem(class_data.get('class_number', ''))
                class_number_item.setTextAlignment(Qt.AlignCenter)
                self.classTable.setItem(row, 2, class_number_item)

                delete_button = QPushButton('删除')
                delete_button.setObjectName('delete-button')
                delete_button.clicked.connect(lambda checked, row=row: self.removeClassRow(row))
                self.classTable.setCellWidget(row, 3, delete_button)

            self.adjustTableHeight(self.classTable)

    def generateResult(self):
        try:
            def convert_to_xlsx(xls_path, xlsx_path):
                df = pd.read_excel(xls_path)
                df.to_excel(xlsx_path, index=False)

            def snake_assign_students(students, class_infos):
                # 将 student_values_slice 按 value 从高到低排序
                sorted_students = sorted(students, key=lambda x: student_values[x], reverse=True)
                
                #students_per_class = len(sorted_students) / len(class_infos)
                
                # 初始化班级索引
                class_index = 0
                direction = 1  # 1 表示向前，-1 表示向后

                for i, name in enumerate(sorted_students):
                    value = student_values[name]
                    gender = gender_data[name]
                    # 当前班级信息
                    current_class_info = class_infos[class_index]

                    # 将学生分配到当前班级
                    class_results[current_class_info].append((name, value, gender))

                    # 更新班级索引
                    if direction == 1:
                        class_index += 1
                        if class_index == len(class_infos):  # 到达末尾，改变方向
                            class_index -= 1
                            direction = -1
                    else:
                        class_index -= 1
                        if class_index < 0:  # 到达开头，改变方向
                            class_index += 1
                            direction = 1

            def balance_gender_ratio(class_result, class_set):
                def get_gender_ratio(class_info):
                    students = class_result[class_info]
                    male_count = sum(1 for student in students if student[2] == '男')
                    total_count = len(students)
                    return male_count / total_count if total_count > 0 else 0
                
                def swap_students(class_info1, class_info2, male_student, female_student):
                    class_result[class_info1].remove(male_student)
                    class_result[class_info1].append(female_student)
                    class_result[class_info2].remove(female_student)
                    class_result[class_info2].append(male_student)

                
                for classes in class_set:
                    exp_c = [class_info for class_info in classes if class_info[1] == '实验班']
                    if exp_c and classes != exp_c:
                        class_set.append(exp_c)
                        classes = [class_info for class_info in classes if class_info[1] == '普通班']
                    total_males = sum(sum(1 for student in class_result[class_info] if student[2] == '男') for class_info in classes)
                    total_students = sum(len(class_result[class_info]) for class_info in classes)
                    target_gender_ratio = total_males / total_students if total_students > 0 else 0
                    
                    classes_to_adjust = classes[:]
                    previous_swaps = set()
                    
                    while classes_to_adjust:
                        flag = True
                        class_gender_ratios = {class_info: get_gender_ratio(class_info) for class_info in classes_to_adjust}
                        for class_info, ratio in list(class_gender_ratios.items()):
                            if abs(ratio - target_gender_ratio) < 0.01:
                                classes_to_adjust.remove(class_info)
                        
                        if not classes_to_adjust:
                            break
                        
                        max_ratio_class = max(classes_to_adjust, key=lambda ci: class_gender_ratios[ci])
                        male_student = next(student for student in class_result[max_ratio_class] if student[2] == '男')
                        
                        potential_females = []
                        for class_info in classes_to_adjust:
                            if class_info != max_ratio_class and get_gender_ratio(class_info) < target_gender_ratio:
                                potential_females.extend([student for student in class_result[class_info] if student[2] == '女'])
                        
                        if potential_females:
                            female_student = min(potential_females, key=lambda student: abs(student[1] - male_student[1]))
                            for class_info in classes_to_adjust:
                                if female_student in class_result[class_info]:
                                    swap_students(max_ratio_class, class_info,
                                                  male_student, female_student)
                                    if (max_ratio_class, class_info) in previous_swaps:
                                        break
                                    previous_swaps.add((class_info, max_ratio_class))
                                    flag = False
                                    break

                        if flag:
                            break
                
                return class_result

            self.progressDialog = ProgressDialog(self)
            self.progressDialog.show()
            # 读取分班表格数据
            self.progressDialog.updateText('读取分班表格数据...')
            class_data = []
            for row in range(self.classTable.rowCount()):
                subject = self.classTable.cellWidget(row, 0).text().split('\n')
                subject = tuple(['、'.join(set(each.split('、'))) for each in subject])
                class_type = self.classTable.cellWidget(row, 1).currentText()
                class_nums = self.classTable.item(row, 2).text().split(',')
                if class_nums == ['']:
                    self.showErrorDialog(f'第 {row+1} 行班级序号未填写')
                    self.closeProgressDialog()
                    return
                for class_num in class_nums:
                    if '-' in class_num:
                        start, end = map(int, class_num.split('-'))
                        class_data.extend([(subject, class_type, i) for i in range(start, end + 1)])
                    else:
                        class_data.append((subject, class_type, int(class_num)))

            # 按照 subject 的元素个数从小到大排序
            class_data = sorted(class_data, key=lambda x: len(x[0]), reverse=False)

            # 读取成绩表格数据
            self.progressDialog.updateText('读取成绩表格数据...')
            folder_path = self.gradeLabel.text().split('：')[-1]
            if not os.path.exists(folder_path):
                self.showErrorDialog('未选择成绩文件夹')
                return
            
            folder_files = os.listdir(folder_path)
            for row in range(self.gradeTable.rowCount()):
                exam_name = self.gradeTable.item(row, 0).text()
                if not exam_name + '.xlsx' in folder_files:
                    convert_to_xlsx(os.path.join(folder_path, exam_name + '.xls'),
                                    os.path.join(folder_path, exam_name + '.xlsx'))

            grade_files = []
            for row in range(self.gradeTable.rowCount()):
                exam_name = self.gradeTable.item(row, 0).text()
                weightText = self.gradeTable.cellWidget(row, 1).text()
                if weightText == '':
                    self.showErrorDialog(f'第 {row+1} 行权重未填写')
                    return
                weight = float(weightText) / 100
                name_col = self.gradeTable.cellWidget(row, 2).text()
                score_col = self.gradeTable.cellWidget(row, 3).text()
                rank_col = self.gradeTable.cellWidget(row, 4).text()
                grade_files.append((exam_name, weight, name_col, score_col, rank_col))

            if not grade_files:
                self.showErrorDialog('成绩文件夹中无成绩表格')
                return

            # 读取选科表格
            self.progressDialog.updateText('读取选科表格数据...')
            sub_path = self.subLabel.text().split('：')[-1]
            if not os.path.exists(sub_path):
                self.showErrorDialog("未选择性别或选科表格")
                return
            if sub_path[-1] != 'x':
                convert_to_xlsx(sub_path, sub_path + 'x')
                sub_path += 'x'
            sub_data = {}
            sub_amount = {}
            if sub_path:
                sub_wb = openpyxl.load_workbook(sub_path)
                sub_ws = sub_wb.active
                for row in sub_ws.iter_rows(min_row=2, values_only=True):
                    name, *subjects = row
                    subjects = '、'.join(set(subjects))
                    sub_data[name] = subjects
                    if sub_amount.get(subjects) == None:
                        sub_amount[subjects] = 0
                    sub_amount[subjects] += 1

            sorted_sub_data = sorted(sub_data.items(), key=lambda x: sub_amount[x[1]],
                                     reverse=True)
            sub_stu_set = set(sub_data.keys())

            # 读取性别表格
            self.progressDialog.updateText('读取性别表格数据...')
            gender_path = self.genderLabel.text().split('：')[-1]
            if not os.path.exists(gender_path):
                self.showErrorDialog("请确保已选择性别和选科表格")
                return
            if gender_path[-1] != 'x':
                convert_to_xlsx(gender_path, gender_path + 'x')
                gender_path += 'x'
            gender_data = {}
            if gender_path:
                gender_wb = openpyxl.load_workbook(gender_path, data_only=True)
                gender_ws = gender_wb.active
                for row in gender_ws.iter_rows(min_row=2, values_only=True):
                    name, gender = row
                    gender_data[name] = gender
            gender_stu_set = set(gender_data.keys())
            
            stus = sub_stu_set & gender_stu_set
            only_gender = [student for student in gender_stu_set if student not in sub_stu_set]
            only_sub = [student for student in sub_stu_set if student not in gender_stu_set]
            if only_gender:
                self.progressDialog.updateText(f'以下学生无选科数据，将被程序忽略：\n{" ".join(only_gender)}\n')
            if only_sub:
                self.progressDialog.updateText(f'以下学生无性别数据，将被程序忽略：\n{" ".join(only_sub)}\n')

            # 计算学生 value 值
            self.progressDialog.updateText('正在分班...')
            student_values = {stu: [] for stu in stus}
            folder_path = self.gradeLabel.text().split('：')[-1]
            not_exist = []
            for exam_name, weight, name_col, score_col, rank_col in grade_files:
                grade_wb = openpyxl.load_workbook(os.path.join(folder_path,
                                                               exam_name + '.xlsx'))
                grade_ws = grade_wb.active
                total_students = grade_ws.max_row - 1
                for row in grade_ws.iter_rows(min_row=2, values_only=True):
                    name = row[ord(name_col) - ord('A')]
                    rank = row[ord(rank_col) - ord('A')]
                    if name not in student_values:
                        #student_values[name] = []
                        not_exist.append(name)
                        continue
                    value = total_students - rank
                    if rank == 0:
                        value = -1
                    student_values[name].append((value, weight))
                if not_exist:
                    self.progressDialog.updateText(f'在 {exam_name} 中，以下学生无选科或性别数据，将被程序忽略：\n{" ".join([student for student in not_exist if student not in only_gender + only_sub])}\n')

            # 计算加权平均 value 值
            no_exam = []
            for name, data in student_values.items():
                valid_values = [each[0] for each in data if each[0] != -1]
                avg_value = valid_values and sum(valid_values) / len(valid_values) or 0
                if avg_value == 0:
                    no_exam.append(name)
                for index, each in enumerate(data):
                    if each[0] == -1:
                        data[index] = (avg_value, each[1])
                student_values[name] = sum([value * weight for value, weight in data])
            self.progressDialog.updateText(f'以下学生未参加任何考试，将以平均最后一名计算：\n{" ".join([student for student in no_exam])}\n')

            #getClassSet
            sub_set = [] #[set, set, set]
            class_set = [] #[list, list, list]
            for class_info in class_data:
                res = -1
                del_list = []
                for index, ss in enumerate(sub_set):
                    if set(class_info[0]) & ss != set():
                        if res == -1:
                            ss |= set(class_info[0])
                            class_set[index].append(class_info)
                            res = index
                        else:
                            sub_set[res] |= ss
                            class_set[res] += class_set[index]
                            del_list.append(index)
                for index in del_list:
                    del sub_set[index]
                    del class_set[index]
                            
                if res == -1:
                    sub_set.append(set(class_info[0]))
                    class_set.append([class_info])

            class_results = {class_info: [] for class_info in class_data}
            sub_stu = {subjects: sorted([student[0] for student in students if student[0] in stus],
                                        key=lambda x: student_values[x], reverse=True) \
                       for subjects, students in groupby(sorted_sub_data, key=lambda x: x[1])}
                # {sub1: [stus], sub2: [stus], ...}
            sub_iter = {sub: 0 for sub in sub_stu}
            for index, classes in enumerate(class_set):
                class_size = sum([len(stu) for sub, stu in sub_stu.items() \
                                 if sub in sub_set[index]]) / len(classes)
                class_size = round(class_size)
                classes = sorted(classes, key=lambda x: (len(x[0]), x[0], x[1] != "实验班"))
                    # 由于布尔值 False（"实验班"）在排序时比 True（"普通班"）更靠前
                    # 因此这个比较将 "实验班" 放在前面
                classes = {sub_and_type: [class_info for class_info in class_infos] \
                           for sub_and_type, class_infos in groupby(classes, key=lambda x: (x[0], x[1]))}
                    #{((subs), type): [class_infos]}
                for sub_and_type, class_infos in classes.items():
                    if len(sub_and_type[0]) == 1:
                        sub = sub_and_type[0][0]
                        amount = class_size * len(class_infos)
                        snake_assign_students(sub_stu[sub][sub_iter[sub] :
                                                           sub_iter[sub] + amount - 1], class_infos)
                        sub_iter[sub] += amount
                    else:
                        for sub in sub_and_type[0]:
                            snake_assign_students(sub_stu[sub][sub_iter[sub] :], class_infos)

            '''
            [print(class_info, sum([value for name, value, gender in stu]) / \
                   len(stu), sum([gender == '男' for name, value, gender in stu]) , \
                                   sum([gender == '女' for name, value, gender in stu])) \
             for class_info, stu in class_results.items()]
             '''

            # 平衡班级性别比例
            self.progressDialog.updateText('平衡班级性别比例...')
            class_results = balance_gender_ratio(class_results, class_set)

            class_results = dict(sorted(class_results.items(), key = lambda x: x[0][2]))
            '''
            [print(class_info, sum([value for name, value, gender in stu]) / len(stu),
             sum([gender == '男' for name, value, gender in stu]) , \
                                   sum([gender == '女' for name, value, gender in stu])) \
             for class_info, stu in class_results.items()]
             '''

            # 输出结果到 Excel
            output_wb = openpyxl.Workbook()
            output_ws = output_wb.active
            output_ws.append(['姓名', '班级', '性别', '班型', '层次'])
            for class_info, students in class_results.items():
                subject, class_type, class_num = class_info
                for name, value, gender in students:
                    output_ws.append([name, f'{class_num}', gender, f'{subject}', class_type])

            now_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            output_wb.save(f'分班结果 {now_time}.xlsx')
            self.progressDialog.updateText('完成！')

            self.showSuccessDialog(f'分班结果已保存到 分班结果 {now_time}.xlsx')
            if not (only_gender or only_sub or not_exist or no_exam):
                self.closeProgressDialog()
        except Exception as e:
            self.showErrorDialog(f'发生错误：\n{traceback.format_exc()}')
            self.closeProgressDialog()

    def showErrorDialog(self, message):
        QMessageBox.critical(self, "错误", message)

    def showSuccessDialog(self, message):
        QMessageBox.information(self, "成功", message)

    def closeProgressDialog(self):
        if hasattr(self, 'progressDialog') and self.progressDialog.isVisible():
            self.progressDialog.close()
                    
if __name__ == "__main__":
    app = QApplication(sys.argv)
    widget = MainWidget()
    widget.show()
    sys.exit(app.exec_())

